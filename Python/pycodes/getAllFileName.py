import os
from openpyxl import Workbook

# 定义文件夹路径
folder_path = "/Users/a1234/Documents/项目上传-上海生物医药创新转化基金"

# 创建一个新的Excel工作簿
wb = Workbook()
ws = wb.active
ws.title = "File Names"
# 写入表头
ws['A1'] = 'File Names'

# 递归遍历文件夹下所有子文件并写入Excel表格
def list_files(start_folder, ws, current_row):
    for root, dirs, files in os.walk(start_folder):
        for file in files:
            file_path = os.path.join(root, file)
            ws.cell(row=current_row, column=1, value=file_path)
            current_row += 1
    return current_row

# 从第2行开始写入文件名
row = 2
row = list_files(folder_path, ws, row)

# 保存Excel表格至指定文件夹
excel_path = os.path.join(folder_path, "file_names.xlsx")
wb.save(excel_path)

print("所有子文件名已整理到 Excel 表格中，并保存至指定文件夹！")